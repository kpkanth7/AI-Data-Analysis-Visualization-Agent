import os
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXPORTS_DIR = "exports"


def _ensure_exports_dir():
    os.makedirs(EXPORTS_DIR, exist_ok=True)


def export_to_excel(df: pd.DataFrame, path: str | None = None, title: str = "Data Export") -> str:
    _ensure_exports_dir()
    if path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join(EXPORTS_DIR, f"export_{ts}.xlsx")

    df.to_excel(path, index=False, engine="openpyxl")

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    ws.title = "Data"

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="EEEEEE"),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    for col_idx, col in enumerate(df.columns, 1):
        max_len = max(
            len(str(col)),
            df[col].astype(str).str.len().max() if len(df) > 0 else 0,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"
    wb.save(path)
    return path


def export_to_csv(df: pd.DataFrame, path: str | None = None) -> str:
    _ensure_exports_dir()
    if path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join(EXPORTS_DIR, f"export_{ts}.csv")
    df.to_csv(path, index=False)
    return path
